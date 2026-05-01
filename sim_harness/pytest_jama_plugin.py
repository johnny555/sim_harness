# Copyright 2026 The sim_harness Authors
# SPDX-License-Identifier: Apache-2.0

"""
Pytest plugin: collect test outcomes and (optionally) export them as a
Jama-compatible Excel spreadsheet.

Auto-loaded via the ``pytest11`` entry point in ``setup.cfg`` whenever
``sim_harness`` is on the import path; no conftest copy-paste needed.

Features
--------
* ``--jama-xlsx PATH``        Write results to ``PATH`` at session end.
                              When omitted, no xlsx is written.
* ``--jama-project KEY``      Project key prefix used in the Item ID
                              column (default: ``DCE``).
* ``@pytest.mark.requirement(req_id, description, category="")``
                              Stamp a Jama requirement ID onto a test.
                              Replaces the old ``assert_requirement`` mixin.

Usage
-----
::

    @pytest.mark.requirement("SIM-SYSRQ_SIM-5", "Maximum 4 vehicles", category="Vehicle")
    def test_quad_vehicle_launch_supports_4_ids():
        ...
        assert ...

    pytest                                    # run; no xlsx output
    pytest --jama-xlsx out.xlsx               # full xlsx export
    pytest -m "requirement"                   # only requirement-tagged tests
"""

from __future__ import annotations

import datetime
import os
import time
from dataclasses import dataclass, field
from typing import List, Optional


# ---------------------------------------------------------------------------
# Pytest hooks
# ---------------------------------------------------------------------------


def pytest_addoption(parser):
    group = parser.getgroup("jama")
    group.addoption(
        "--jama-xlsx",
        default=None,
        help="Write a Jama-compatible Excel report to PATH at session end. "
             "If omitted, no spreadsheet is written.",
    )
    group.addoption(
        "--jama-project",
        default="DCE",
        help="Project key prefix used in the Item ID column (default: DCE).",
    )


def pytest_configure(config):
    config.addinivalue_line(
        "markers",
        "requirement(req_id, description, category=''): "
        "tag a test with a Jama requirement ID for traceability.",
    )
    collector = JamaResultCollector(config)
    config._jama_collector = collector
    config.pluginmanager.register(collector, "jama_xlsx_exporter")


# ---------------------------------------------------------------------------
# Result collection
# ---------------------------------------------------------------------------


@dataclass
class _Result:
    nodeid: str
    test_name: str
    status: str             # PASSED / FAILED / NOT_RUN / BLOCKED
    duration_s: float
    message: str
    when: str               # 'call' | 'setup' | 'teardown'
    requirement_id: str = ""
    requirement_description: str = ""
    requirement_category: str = ""


class JamaResultCollector:
    """Pytest plugin that collects outcomes and writes a Jama-importable xlsx."""

    def __init__(self, config):
        self.config = config
        self.results: List[_Result] = []
        self._start_times = {}
        self._req_marks = {}  # nodeid -> (req_id, description, category)

    # ---- result capture ----------------------------------------------------

    def pytest_runtest_logstart(self, nodeid, location):
        self._start_times[nodeid] = time.time()

    def pytest_runtest_setup(self, item):
        # Stash the requirement marker (if any) so we can attach it to the
        # eventual report. Markers are scoped to the test item and aren't
        # reachable from the report object alone.
        marker = item.get_closest_marker("requirement")
        if marker is not None:
            req_id = marker.args[0] if marker.args else marker.kwargs.get("req_id", "")
            description = (
                marker.args[1] if len(marker.args) > 1
                else marker.kwargs.get("description", "")
            )
            category = marker.kwargs.get("category", "")
            self._req_marks[item.nodeid] = (req_id, description, category)

    def pytest_runtest_logreport(self, report):
        # Only the 'call' phase is interesting unless a setup/teardown failed.
        if not (report.when == "call" or (report.when != "call" and report.failed)):
            return
        duration = time.time() - self._start_times.get(report.nodeid, time.time())
        req_id, req_desc, req_cat = self._req_marks.get(report.nodeid, ("", "", ""))
        self.results.append(
            _Result(
                nodeid=report.nodeid,
                test_name=_friendly_name(report.nodeid),
                status=_map_status(report),
                duration_s=round(duration, 1),
                message=_failure_message(report) if not report.passed else "",
                when=report.when,
                requirement_id=req_id,
                requirement_description=req_desc,
                requirement_category=req_cat,
            )
        )

    # ---- session output ----------------------------------------------------

    def pytest_sessionfinish(self, session, exitstatus):
        xlsx_path = self.config.getoption("jama_xlsx") or os.environ.get("PYTEST_JAMA_XLSX")
        if not xlsx_path or not self.results:
            return
        project_key = self.config.getoption("jama_project") or os.environ.get(
            "PYTEST_JAMA_PROJECT", "DCE")
        try:
            _write_jama_xlsx(self.results, xlsx_path, project_key)
            session.config.get_terminal_writer().write(
                f"\nJama results exported to: {xlsx_path}\n", green=True,
            )
        except ImportError as exc:
            # openpyxl is declared as an exec_depend in sim_harness/package.xml
            # and as install_requires in setup.py, so this should normally be
            # satisfied automatically. If you hit this, run one of:
            #   rosdep install --from-paths src --ignore-src -r -y      # ROS workspace
            #   sudo apt install python3-openpyxl                       # apt
            #   pip install --user --break-system-packages openpyxl     # last resort
            session.config.get_terminal_writer().write(
                f"\nSkipping Jama export — {exc}. "
                f"Run `rosdep install --from-paths src --ignore-src -r -y` "
                f"or `sudo apt install python3-openpyxl` to enable the xlsx writer.\n",
                yellow=True,
            )
        except Exception as exc:  # noqa: BLE001
            # Best-effort: never fail the test session because of an export
            # problem (disk full, permission denied, malformed result row, ...).
            session.config.get_terminal_writer().write(
                f"\nJama export failed ({type(exc).__name__}: {exc}). "
                f"Test results are unaffected.\n",
                yellow=True,
            )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _map_status(report) -> str:
    if report.passed:
        return "PASSED"
    if report.failed:
        return "FAILED"
    if report.skipped:
        return "NOT_RUN"
    return "BLOCKED"


def _friendly_name(nodeid: str) -> str:
    """``test/test_smoke.py::TestX::test_foo`` → ``test_foo``."""
    return nodeid.rsplit("::", 1)[-1]


def _failure_message(report) -> str:
    text = getattr(report, "longreprtext", "") or str(report.outcome)
    if len(text) > 2000:
        text = text[:2000] + "\n... (truncated)"
    return text


def _write_jama_xlsx(results: List[_Result], path: str, project_key: str) -> None:
    """Write or append to a Jama Connect Excel-import-compatible workbook."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    headers = [
        "Item ID",
        "Requirement ID",
        "Name",
        "Description",
        "Status",
        "Failure / Notes",
        "Category",
        "Execution Date",
        "Duration (s)",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    execution_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Ensure parent directory exists before reading/writing the workbook.
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb["Test Results"] if "Test Results" in wb.sheetnames else wb.active
        if ws.max_row < 1:
            ws.title = "Test Results"
            ws.append(headers)
        existing_count = max(ws.max_row - 1, 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"
        ws.append(headers)
        existing_count = 0

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(results, start=existing_count + 1):
        item_id = f"{project_key}-TR-{i:04d}"
        ws.append([
            item_id,
            r.requirement_id,
            r.test_name,
            r.requirement_description,
            r.status,
            r.message,
            r.requirement_category,
            execution_date,
            r.duration_s,
        ])

    status_fills = {
        "PASSED":  PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "FAILED":  PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "NOT_RUN": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "BLOCKED": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    }
    status_col = headers.index("Status") + 1  # 1-based
    for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
        for cell in row:
            if cell.value in status_fills:
                cell.fill = status_fills[cell.value]

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max((min(len(str(c.value)), 60) for c in col if c.value), default=10)
        ws.column_dimensions[col_letter].width = max_len + 4

    # Summary sheet
    if "Summary" in wb.sheetnames:
        ws2 = wb["Summary"]
        ws2.delete_rows(1, ws2.max_row)
    else:
        ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    statuses = [row[4] for row in all_rows if len(row) > 4]
    total = len(all_rows)
    passed = sum(1 for status in statuses if status == "PASSED")
    failed = sum(1 for status in statuses if status == "FAILED")
    skipped = sum(1 for status in statuses if status == "NOT_RUN")
    blocked = sum(1 for status in statuses if status == "BLOCKED")
    tagged = sum(1 for row in all_rows if len(row) > 1 and row[1])
    total_duration = 0.0
    for row in all_rows:
        if len(row) > 8 and row[8] is not None:
            try:
                total_duration += float(row[8])
            except (TypeError, ValueError):
                pass

    ws2.append(["Project", project_key])
    ws2.append(["Execution Date", execution_date])
    ws2.append(["Total Tests", total])
    ws2.append(["Passed", passed])
    ws2.append(["Failed", failed])
    ws2.append(["Skipped / Not Run", skipped])
    ws2.append(["Blocked", blocked])
    ws2.append(["Pass Rate", f"{passed / total * 100:.1f}%" if total else "N/A"])
    ws2.append(["Tagged with Requirement ID", tagged])
    ws2.append(["Total Duration (s)", round(total_duration, 1)])

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 28

    wb.save(path)
